import os
import tempfile
import unittest

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:  # pragma: no cover - зависит от окружения
    Workbook = load_workbook = None  # type: ignore[assignment]

from section3_template import fill_section3_template


@unittest.skipIf(Workbook is None, "openpyxl не установлен")
class Section3TemplateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)

    def _create_template(self, filename: str = "template.xlsx") -> str:
        path = os.path.join(self.temp_dir.name, filename)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Справка"

        headers = (
            "№",
            "Условный номер помещения",
            "Назначение помещения",
            "Этаж расположения",
            "Номер подъезда",
            "Площадь, м2",
            "Наименование помещения",
            "Площадь части, м2",
            "Высота потолков, м",
        )
        for column, value in enumerate(headers, start=2):
            sheet.cell(row=8, column=column).value = value

        # Добавим строку итога ниже предполагаемой таблицы, чтобы убедиться,
        # что функция не затронет её.
        sheet.cell(row=20, column=3).value = "Итого"
        sheet.cell(row=20, column=6).value = "=SUM(F9:F19)"

        workbook.save(path)
        return path

    def test_fill_template_writes_rows_and_numbers(self):
        template_path = self._create_template()
        output_path = os.path.join(self.temp_dir.name, "result.xlsx")

        rows = [
            {
                "unit_number": "НП-1",
                "purpose": "Магазин",
                "floor": "1",
                "entrance_number": "2",
                "area": 35.7,
                "room_name": "Помещение 101",
                "part_area": 35.7,
                "ceiling_height": 3.45,
            },
            {
                "unit_number": "НП-2",
                "purpose": "Кафе",
                "floor": "1",
                "entrance_number": "1",
                "area": 42.1,
                "room_name": "Помещение 102",
                "part_area": 41.3,
                "ceiling_height": 3.6,
            },
        ]

        fill_section3_template(template_path, rows, output_path=output_path)

        workbook = load_workbook(output_path, data_only=True)
        sheet = workbook.active

        self.assertEqual(sheet["B9"].value, 1)
        self.assertEqual(sheet["B10"].value, 2)
        self.assertEqual(sheet["C9"].value, "НП-1")
        self.assertEqual(sheet["D10"].value, "Кафе")
        self.assertEqual(sheet["G9"].value, 35.7)
        self.assertEqual(sheet["I10"].value, 41.3)
        # Строка "Итого" должна сохраниться.
        self.assertEqual(sheet["C20"].value, "Итого")

    def test_fill_template_clears_previous_rows(self):
        template_path = self._create_template()

        workbook = load_workbook(template_path)
        sheet = workbook.active
        sheet["B9"].value = 1
        sheet["C9"].value = "Старые данные"
        sheet["B10"].value = 2
        sheet["C10"].value = "Будет очищено"
        workbook.save(template_path)

        rows = [
            {
                "unit_number": "НП-3",
                "purpose": "Офис",
                "floor": "2",
                "entrance_number": "1",
                "area": 18.5,
                "room_name": "Помещение 201",
                "part_area": 18.5,
                "ceiling_height": 3.2,
            }
        ]

        fill_section3_template(template_path, rows)

        workbook = load_workbook(template_path)
        sheet = workbook.active

        self.assertIsNone(sheet["B10"].value)
        self.assertIsNone(sheet["C10"].value)
        self.assertEqual(sheet["C9"].value, "НП-3")

    def test_fill_template_with_explicit_start_cell(self):
        template_path = self._create_template()
        workbook = load_workbook(template_path)
        sheet = workbook.active
        sheet["J5"].value = "Условный номер помещения"
        sheet["K5"].value = "Назначение помещения"
        sheet["L5"].value = "Этаж расположения"
        sheet["M5"].value = "Номер подъезда"
        sheet["N5"].value = "Площадь, м2"
        sheet["O5"].value = "Наименование помещения"
        sheet["P5"].value = "Площадь части, м2"
        sheet["Q5"].value = "Высота потолков, м"
        workbook.save(template_path)

        rows = [
            {
                "unit_number": "НП-4",
                "purpose": "Фитнес",
                "floor": "1",
                "entrance_number": "3",
                "area": 55.0,
                "room_name": "Помещение 301",
                "part_area": 55.0,
                "ceiling_height": 4.1,
            }
        ]

        fill_section3_template(
            template_path,
            rows,
            data_start_cell="J6",
            numbering_column="I",
        )

        workbook = load_workbook(template_path)
        sheet = workbook.active

        self.assertEqual(sheet["I6"].value, 1)
        self.assertEqual(sheet["J6"].value, "НП-4")
        self.assertEqual(sheet["Q6"].value, 4.1)


if __name__ == "__main__":
    unittest.main()

