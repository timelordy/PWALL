"""Утилиты для заполнения шаблона раздела 3 в формате Excel."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from functools import lru_cache
from typing import Callable, Iterable, Mapping, MutableMapping, Optional, Sequence, TYPE_CHECKING

from section3_parser import SECTION3_FIELDS, identify_section3_field

if TYPE_CHECKING:  # pragma: no cover - используется только для типов
    from openpyxl.worksheet.worksheet import Worksheet

_STOP_KEYWORDS = ("итого", "раздел", "примечан")


def fill_section3_template(
    template_path: str,
    rows: Iterable[Mapping[str, object] | object],
    *,
    output_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    data_start_cell: Optional[str] = None,
    numbering_column: Optional[str | int] = None,
) -> str:
    """Заполняет таблицу раздела 3 в Excel-шаблоне.

    Parameters
    ----------
    template_path:
        Путь до исходного шаблона Excel (``.xlsx``).  Стиль и форматирование
        берутся именно из этого файла.
    rows:
        Последовательность записей раздела 3.  Как правило, это результат
        :func:`section3_parser.extract_section3_rows`.  Допускаются словари,
        объекты с методом ``dict`` или датаклассы — данные будут приведены
        к словарю автоматически.
    output_path:
        Если указан, заполненный файл сохраняется по этому пути.  В
        противном случае исходный ``template_path`` будет перезаписан.
    sheet_name:
        Имя листа, на котором находится таблица.  Если не задано, будет
        использован активный лист книги.
    data_start_cell:
        Ячейка, с которой начинается первая строка таблицы.  Если параметр
        опущен, функция попытается автоматически найти строку с заголовками
        и определить расположение колонок.
    numbering_column:
        Буквенный адрес или индекс колонки с порядковым номером строк.
        При ``None`` попытка определить колонку производится автоматически
        (по заголовку, содержащему символ ``№``).

    Returns
    -------
    str
        Путь к сохранённому файлу (``output_path`` или ``template_path``).
    """

    load_workbook, column_index_from_string, coordinate_from_string = _require_openpyxl()

    workbook = load_workbook(template_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    normalized_rows = [_coerce_row(row) for row in rows]

    header_row, column_map = _resolve_layout(
        worksheet, data_start_cell, column_index_from_string, coordinate_from_string
    )
    data_row_index = (
        coordinate_from_string(data_start_cell)[1]
        if data_start_cell
        else header_row + 1
    )

    numbering_index = _resolve_numbering_column(
        worksheet, numbering_column, header_row, column_map, column_index_from_string
    )

    existing_rows = _count_existing_rows(worksheet, column_map, data_row_index)
    rows_to_clear = max(existing_rows, len(normalized_rows))
    if not rows_to_clear:
        rows_to_clear = 1

    _clear_rows(worksheet, column_map, data_row_index, rows_to_clear, numbering_index)

    for offset, row in enumerate(normalized_rows):
        current_row = data_row_index + offset
        if numbering_index is not None:
            worksheet.cell(row=current_row, column=numbering_index).value = offset + 1
        for field, column_index in column_map.items():
            worksheet.cell(row=current_row, column=column_index).value = row.get(field)

    destination = output_path or template_path
    workbook.save(destination)
    return destination


def _coerce_row(row: Mapping[str, object] | object) -> MutableMapping[str, object]:
    if isinstance(row, Mapping):
        return {str(key): value for key, value in row.items()}

    if is_dataclass(row):
        return dict(asdict(row))

    to_dict = getattr(row, "to_dict", None)
    if callable(to_dict):
        result = to_dict()
        if isinstance(result, Mapping):
            return {str(key): value for key, value in result.items()}

    raise TypeError("row must be mapping-like or convertible to dict")


def _resolve_layout(
    worksheet: "Worksheet",
    data_start_cell: Optional[str],
    column_index_from_string: Callable[[str], int],
    coordinate_from_string: Callable[[str], tuple[str, int]],
) -> tuple[int, dict[str, int]]:
    if data_start_cell:
        column_letter, row_index = coordinate_from_string(data_start_cell)
        start_column = column_index_from_string(column_letter)
        column_map = {
            field: start_column + offset for offset, field in enumerate(SECTION3_FIELDS)
        }
        return row_index - 1, column_map

    for row in worksheet.iter_rows():
        header_row_index = row[0].row
        mapping: dict[str, int] = {}
        for cell in row:
            field = identify_section3_field(cell.value)
            if field and field not in mapping:
                mapping[field] = cell.col_idx
        if len(mapping) == len(SECTION3_FIELDS):
            return header_row_index, mapping

    raise ValueError("Не удалось найти строку с заголовками раздела 3")


def _resolve_numbering_column(
    worksheet: "Worksheet",
    numbering_column: Optional[str | int],
    header_row: int,
    column_map: Mapping[str, int],
    column_index_from_string: Callable[[str], int],
) -> Optional[int]:
    if numbering_column is None:
        left_column = min(column_map.values()) - 1
        if left_column >= 1:
            header_value = worksheet.cell(row=header_row, column=left_column).value
            if isinstance(header_value, str) and "№" in header_value:
                return left_column
        return None

    if isinstance(numbering_column, int):
        if numbering_column < 1:
            raise ValueError("numbering_column must be positive")
        return numbering_column

    column = numbering_column.strip()
    if not column:
        return None
    return column_index_from_string(column)


def _count_existing_rows(
    worksheet: "Worksheet",
    column_map: Mapping[str, int],
    start_row: int,
) -> int:
    max_row = worksheet.max_row
    row_index = start_row
    count = 0
    while row_index <= max_row:
        values = [worksheet.cell(row=row_index, column=col).value for col in column_map.values()]
        if _is_footer_row(values):
            break
        if all(_is_empty(value) for value in values):
            break
        count += 1
        row_index += 1
    return count


def _clear_rows(
    worksheet: "Worksheet",
    column_map: Mapping[str, int],
    start_row: int,
    count: int,
    numbering_column: Optional[int],
) -> None:
    for offset in range(count):
        row_index = start_row + offset
        for column_index in column_map.values():
            worksheet.cell(row=row_index, column=column_index).value = None
        if numbering_column is not None:
            worksheet.cell(row=row_index, column=numbering_column).value = None


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _is_footer_row(values: Sequence[object]) -> bool:
    for value in values:
        if isinstance(value, str):
            normalized = value.strip().lower()
            if any(keyword in normalized for keyword in _STOP_KEYWORDS):
                return True
    return False


__all__ = ["fill_section3_template"]


@lru_cache(maxsize=1)
def _require_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore import
        from openpyxl.utils.cell import (  # type: ignore import
            column_index_from_string,
            coordinate_from_string,
        )
    except ModuleNotFoundError as exc:  # pragma: no cover - зависит от окружения
        raise ModuleNotFoundError(
            "Для заполнения шаблона необходим пакет openpyxl. "
            "Установите его командой 'pip install openpyxl'."
        ) from exc

    return load_workbook, column_index_from_string, coordinate_from_string

